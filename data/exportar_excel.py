import sqlite3
import os
from datetime import datetime
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv()

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, LineChart
except ImportError:
    print("Instale openpyxl: pip install openpyxl")
    exit()

DB_PATH = os.path.join(os.path.dirname(__file__), "edge_protocol.db")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "logs", "edge_protocol_resultados.xlsx")

COR_VERDE        = "C6EFCE"
COR_VERDE_TEXTO  = "276221"
COR_VERMELHO     = "FFC7CE"
COR_VERM_TEXTO   = "9C0006"
COR_PENDENTE     = "FFEB9C"
COR_PEND_TEXTO   = "9C5700"
COR_HEADER       = "1F4E79"
COR_SUBHEADER    = "2E75B6"
COR_ALTERNADO    = "DEEAF1"
COR_BRANCO       = "FFFFFF"
COR_CINZA_CLARO  = "F2F2F2"
COR_DESTAQUE     = "00B0F0"
COR_OURO         = "FFD700"

def buscar_todos_sinais():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT data, liga, jogo, mercado, odd, ev_estimado,
               edge_score, stake_unidades, status, resultado,
               lucro_unidades, criado_em
        FROM sinais ORDER BY criado_em ASC
    ''')
    rows = c.fetchall()
    conn.close()
    return rows

def calcular_resumo(sinais):
    total = len(sinais)
    vitorias = sum(1 for s in sinais if s[9] == "verde")
    derrotas = sum(1 for s in sinais if s[9] == "vermelho")
    pendentes = sum(1 for s in sinais if s[8] == "pendente")
    lucro = sum(s[10] for s in sinais if s[10] is not None)
    win_rate = (vitorias / (vitorias + derrotas) * 100) if (vitorias + derrotas) > 0 else 0
    banca_inicial = 1000
    unidade = banca_inicial * 0.01
    banca_atual = banca_inicial + (lucro * unidade)
    roi = ((banca_atual - banca_inicial) / banca_inicial) * 100
    return {
        "total": total, "vitorias": vitorias, "derrotas": derrotas,
        "pendentes": pendentes, "lucro": lucro, "win_rate": win_rate,
        "banca_atual": banca_atual, "roi": roi
    }

def mercado_legivel(mercado):
    mapa = {
        "1x2_casa": "Vitória Casa", "1x2_fora": "Vitória Fora",
        "over_2.5": "Over 2.5", "under_2.5": "Under 2.5",
        "btts_sim": "Ambas Marcam", "btts_nao": "Nenhuma Marca",
    }
    return mapa.get(mercado, mercado)

def borda_fina():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def borda_media():
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, row, col, valor="", bold=False, cor_fundo=None, align="center",
        cor_fonte="000000", tamanho=10, borda=None, wrap=False):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, color=cor_fonte, size=tamanho, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if borda:
        c.border = borda
    if cor_fundo:
        c.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
    return c

def merge_cel(ws, row, col_ini, col_fim, valor="", bold=False, cor_fundo=None,
              align="center", cor_fonte="000000", tamanho=11):
    ws.merge_cells(start_row=row, start_column=col_ini, end_row=row, end_column=col_fim)
    c = ws.cell(row=row, column=col_ini, value=valor)
    c.font = Font(bold=bold, color=cor_fonte, size=tamanho, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center")
    if cor_fundo:
        c.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
    return c

def criar_dashboard(wb, sinais, resumo):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    for col in ["B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws.column_dimensions[col].width = 14
    for row in range(1, 60):
        ws.row_dimensions[row].height = 22

    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 10
    ws.row_dimensions[9].height = 10
    ws.row_dimensions[15].height = 10
    ws.row_dimensions[21].height = 10

    merge_cel(ws, 2, 2, 12, "⚡  EDGE PROTOCOL — DASHBOARD DE PERFORMANCE",
              bold=True, cor_fundo=COR_HEADER, cor_fonte="FFFFFF", tamanho=16, align="center")

    merge_cel(ws, 3, 2, 12, f"Atualizado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
              cor_fundo=COR_SUBHEADER, cor_fonte="FFFFFF", tamanho=9)

    cards = [
        ("TOTAL DE SINAIS", str(resumo["total"]), "2E75B6"),
        ("VITÓRIAS", str(resumo["vitorias"]), "375623"),
        ("DERROTAS", str(resumo["derrotas"]), "9C0006"),
        ("WIN RATE", f"{resumo['win_rate']:.1f}%", "7030A0"),
        ("ROI TOTAL", f"{resumo['roi']:.2f}%", "C55A11" if resumo["roi"] < 0 else "375623"),
        ("BANCA ATUAL", f"R${resumo['banca_atual']:.2f}", "1F4E79"),
    ]

    posicoes = [2, 4, 6, 8, 10, 12]
    for i, (titulo, valor, cor) in enumerate(cards):
        col = posicoes[i] if i < 3 else posicoes[i]
        row_titulo = 5 if i < 3 else 11
        row_valor = 6 if i < 3 else 12
        row_borda = 7 if i < 3 else 13
        col_start = 2 + (i % 3) * 2 if True else 2
        col_start = [2, 5, 8, 2, 5, 8][i]
        col_end = col_start + 1

        ws.merge_cells(start_row=row_titulo, start_column=col_start,
                       end_row=row_titulo, end_column=col_end)
        c = ws.cell(row=row_titulo, column=col_start, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=row_valor, start_column=col_start,
                       end_row=row_valor, end_column=col_end)
        c = ws.cell(row=row_valor, column=col_start, value=valor)
        c.font = Font(bold=True, color=cor, size=18, name="Calibri")
        c.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            left=Side(style="medium", color=cor),
            right=Side(style="medium", color=cor),
            bottom=Side(style="medium", color=cor)
        )

        ws.merge_cells(start_row=row_borda, start_column=col_start,
                       end_row=row_borda, end_column=col_end)
        c = ws.cell(row=row_borda, column=col_start)
        c.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        ws.row_dimensions[row_borda].height = 4

    merge_cel(ws, 17, 2, 12, "PERFORMANCE POR LIGA",
              bold=True, cor_fundo=COR_HEADER, cor_fonte="FFFFFF", tamanho=11)

    ligas = defaultdict(lambda: {"total": 0, "v": 0, "d": 0, "lucro": 0})
    for s in sinais:
        liga = s[1]
        ligas[liga]["total"] += 1
        if s[9] == "verde":
            ligas[liga]["v"] += 1
        elif s[9] == "vermelho":
            ligas[liga]["d"] += 1
        if s[10]:
            ligas[liga]["lucro"] += s[10]

    headers_liga = ["Liga", "Sinais", "Vitórias", "Derrotas", "Win Rate", "Lucro (u)"]
    for col, h in enumerate(headers_liga, 2):
        cel(ws, 18, col, h, bold=True, cor_fundo=COR_SUBHEADER,
            cor_fonte="FFFFFF", borda=borda_fina(), tamanho=10)

    for i, (liga, dados) in enumerate(ligas.items()):
        row = 19 + i
        ws.row_dimensions[row].height = 20
        fin = dados["v"] + dados["d"]
        wr = (dados["v"] / fin * 100) if fin > 0 else 0
        cor_row = COR_ALTERNADO if i % 2 == 0 else COR_BRANCO
        valores = [liga, dados["total"], dados["v"], dados["d"],
                   f"{wr:.0f}%", f"{dados['lucro']:+.2f}u"]
        for col, v in enumerate(valores, 2):
            cel(ws, row, col, v, cor_fundo=cor_row, borda=borda_fina())

    merge_cel(ws, 26, 2, 12, "PERFORMANCE POR MERCADO",
              bold=True, cor_fundo=COR_HEADER, cor_fonte="FFFFFF", tamanho=11)

    mercados = defaultdict(lambda: {"total": 0, "v": 0, "d": 0, "lucro": 0})
    for s in sinais:
        m = mercado_legivel(s[3])
        mercados[m]["total"] += 1
        if s[9] == "verde":
            mercados[m]["v"] += 1
        elif s[9] == "vermelho":
            mercados[m]["d"] += 1
        if s[10]:
            mercados[m]["lucro"] += s[10]

    headers_merc = ["Mercado", "Sinais", "Vitórias", "Derrotas", "Win Rate", "Lucro (u)"]
    for col, h in enumerate(headers_merc, 2):
        cel(ws, 27, col, h, bold=True, cor_fundo=COR_SUBHEADER,
            cor_fonte="FFFFFF", borda=borda_fina(), tamanho=10)

    for i, (merc, dados) in enumerate(mercados.items()):
        row = 28 + i
        ws.row_dimensions[row].height = 20
        fin = dados["v"] + dados["d"]
        wr = (dados["v"] / fin * 100) if fin > 0 else 0
        cor_row = COR_ALTERNADO if i % 2 == 0 else COR_BRANCO
        valores = [merc, dados["total"], dados["v"], dados["d"],
                   f"{wr:.0f}%", f"{dados['lucro']:+.2f}u"]
        for col, v in enumerate(valores, 2):
            cel(ws, row, col, v, cor_fundo=cor_row, borda=borda_fina())

def criar_aba_sinais(wb, sinais):
    ws = wb.create_sheet("Sinais Completos")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.freeze_panes = "B2"

    headers = ["Data", "Liga", "Jogo", "Mercado", "Odd", "EV",
               "Score", "Stake", "Resultado", "Lucro (u)", "ROI Acum."]
    larguras = [12, 18, 35, 16, 8, 10, 8, 8, 12, 12, 12]

    ws.row_dimensions[1].height = 28
    for col, (h, w) in enumerate(zip(headers, larguras), 2):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda_fina()

    lucro_acumulado = 0
    for i, s in enumerate(sinais):
        row = i + 2
        ws.row_dimensions[row].height = 20
        data, liga, jogo, mercado, odd, ev, score, stake, status, resultado, lucro, _ = s

        if lucro:
            lucro_acumulado += lucro

        if resultado == "verde":
            cor = COR_VERDE
            cor_txt = COR_VERDE_TEXTO
            res_txt = "✅ VERDE"
        elif resultado == "vermelho":
            cor = COR_VERMELHO
            cor_txt = COR_VERM_TEXTO
            res_txt = "❌ VERMELHO"
        else:
            cor = COR_PENDENTE
            cor_txt = COR_PEND_TEXTO
            res_txt = "⏳ PENDENTE"

        ev_fmt = f"{ev*100:.2f}%" if ev else "-"
        lucro_fmt = f"{lucro:+.2f}" if lucro is not None else "-"
        roi_fmt = f"{(lucro_acumulado * 10):.2f}%" if lucro_acumulado != 0 else "0.00%"

        valores = [data, liga, jogo, mercado_legivel(mercado), odd,
                   ev_fmt, score, f"{stake}u", res_txt, lucro_fmt, roi_fmt]

        for col, valor in enumerate(valores, 2):
            c = ws.cell(row=row, column=col, value=valor)
            c.font = Font(color=cor_txt if col == 10 else "000000", size=10, name="Calibri")
            c.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            c.alignment = Alignment(
                horizontal="left" if col == 4 else "center",
                vertical="center"
            )
            c.border = borda_fina()

    ws.auto_filter.ref = f"B1:{get_column_letter(len(headers)+1)}1"

def criar_aba_mensal(wb, sinais, mes_ano):
    nome_aba = f"Mês {mes_ano}"
    if nome_aba in [s.title for s in wb.worksheets]:
        return

    ws = wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.freeze_panes = "B2"

    sinais_mes = [s for s in sinais if s[0].startswith(mes_ano)]
    resumo_mes = calcular_resumo(sinais_mes)

    ws.row_dimensions[1].height = 35
    ws.merge_cells("B1:L1")
    c = ws.cell(row=1, column=2, value=f"⚡ EDGE PROTOCOL — {mes_ano}")
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")

    cards_mes = [
        ("Sinais", resumo_mes["total"]),
        ("Vitórias", resumo_mes["vitorias"]),
        ("Derrotas", resumo_mes["derrotas"]),
        ("Win Rate", f"{resumo_mes['win_rate']:.1f}%"),
        ("ROI", f"{resumo_mes['roi']:.2f}%"),
        ("Banca", f"R${resumo_mes['banca_atual']:.2f}"),
    ]

    for i, (titulo, valor) in enumerate(cards_mes):
        col = 2 + i * 2
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        c = ws.cell(row=3, column=col, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.fill = PatternFill(start_color=COR_SUBHEADER, end_color=COR_SUBHEADER, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=4, column=col, value=valor)
        c.font = Font(bold=True, color=COR_HEADER, size=13, name="Calibri")
        c.fill = PatternFill(start_color=COR_ALTERNADO, end_color=COR_ALTERNADO, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda_fina()
        ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions[get_column_letter(col+1)].width = 12

    headers = ["Data", "Liga", "Jogo", "Mercado", "Odd", "EV",
               "Score", "Stake", "Resultado", "Lucro (u)"]
    larguras = [12, 18, 35, 16, 8, 10, 8, 8, 14, 12]

    ws.row_dimensions[6].height = 26
    for col, (h, w) in enumerate(zip(headers, larguras), 2):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=6, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda_fina()

    for i, s in enumerate(sinais_mes):
        row = i + 7
        ws.row_dimensions[row].height = 20
        data, liga, jogo, mercado, odd, ev, score, stake, status, resultado, lucro, _ = s

        if resultado == "verde":
            cor = COR_VERDE
            res_txt = "✅ VERDE"
        elif resultado == "vermelho":
            cor = COR_VERMELHO
            res_txt = "❌ VERMELHO"
        else:
            cor = COR_PENDENTE
            res_txt = "⏳ PENDENTE"

        ev_fmt = f"{ev*100:.2f}%" if ev else "-"
        lucro_fmt = f"{lucro:+.2f}" if lucro is not None else "-"

        valores = [data, liga, jogo, mercado_legivel(mercado), odd,
                   ev_fmt, score, f"{stake}u", res_txt, lucro_fmt]

        for col, valor in enumerate(valores, 2):
            c = ws.cell(row=row, column=col, value=valor)
            c.font = Font(size=10, name="Calibri")
            c.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            c.alignment = Alignment(
                horizontal="left" if col == 4 else "center",
                vertical="center"
            )
            c.border = borda_fina()

def gerar_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    sinais = buscar_todos_sinais()
    resumo = calcular_resumo(sinais)

    wb = openpyxl.Workbook()

    criar_dashboard(wb, sinais, resumo)
    criar_aba_sinais(wb, sinais)

    meses = sorted(set(s[0][:7] for s in sinais), reverse=True)
    for mes in meses:
        criar_aba_mensal(wb, sinais, mes)

    wb.save(EXCEL_PATH)
    print(f"Excel gerado: {EXCEL_PATH}")
    return EXCEL_PATH

if __name__ == "__main__":
    print("=== GERANDO RELATÓRIO EXCEL ===\n")
    gerar_excel()
    print("Concluído!")
"""
update_excel.py — Edge Protocol
Atualiza o arquivo bot_apostas.xlsx após cada aposta.

USO:
    python update_excel.py '{"acao":"nova_aposta","aposta":{...}}'
    python update_excel.py '{"acao":"resultado","aposta":{...}}'
    python update_excel.py '{"acao":"full_refresh"}'

AÇÕES:
    nova_aposta   — insere nova linha na aba Sinais Completos
    resultado     — atualiza resultado, CLV e Brier de uma aposta
    atualizar_odds — atualiza odds na aba Monitoramento 72h
    full_refresh   — regenera o arquivo completo com todos os dados
"""

import sys
import json
import os
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── CONFIG ────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot_apostas.xlsx")
LOG_PATH   = "update_log.txt"

# ── PALETA ────────────────────────────────────────────────────
AZUL      = "1a3a5c"
AZUL_MED  = "2E75B6"
AZUL_CLR  = "DEEAF1"
VERDE_CLR = "C6EFCE"
VERDE_TXT = "276221"
VERM_CLR  = "FFC7CE"
VERM_TXT  = "9C0006"
AMAR_CLR  = "FFEB9C"
AMAR_TXT  = "9C5700"
CINZA     = "f8f8f8"
BRANCO    = "FFFFFF"

def F(cor):
    return PatternFill(start_color=cor, end_color=cor, fill_type="solid")

def bf():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{ts}] {msg}"
    print(linha)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(linha + "\n")

# ── HELPERS ───────────────────────────────────────────────────
def cor_resultado(v):
    if v == "Green":   return VERDE_CLR, VERDE_TXT
    elif v == "Red":   return VERM_CLR,  VERM_TXT
    elif v == "Void":  return "E2EFDA",  "375623"
    return AMAR_CLR, AMAR_TXT  # Pendente

def formatar_celula(c, v, bold=False, bg=None, fg="000000",
                    sz=10, h="center", border=True):
    c.value = v
    c.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
    c.alignment = Alignment(horizontal=h, vertical="center")
    if bg:
        c.fill = F(bg)
    if border:
        c.border = bf()

# ── MONTA LINHA DA APOSTA ─────────────────────────────────────
def linha_aposta(a):
    clv = None
    if a.get("odd_fechamento") and a.get("odd_entrada"):
        clv = ((a["odd_fechamento"]/a["odd_entrada"])-1)*100

    return [
        a.get("id",""),
        a.get("data",""),
        a.get("hora",""),
        a.get("liga",""),
        a.get("jogo",""),
        a.get("mercado",""),
        a.get("tip",""),
        a.get("odd_entrada",""),
        a.get("odd_fechamento") or "-",
        f"{clv:+.2f}%" if clv is not None else "-",
        f"{a.get('ev',0):.1f}%",
        a.get("edge_score",""),
        a.get("tier",""),
        f"{a.get('prob_modelo',0)*100:.0f}%",
        str(a.get("confianca","")),
        f"{a.get('steam',0):+.1f}%",
        f"{a.get('kelly_pct',0):.1f}%",
        f"{a.get('unidades',0):.2f}u",
        a.get("resultado") or "Pendente",
        f"{a['retorno']:+.2f}u" if a.get("retorno") is not None else "-",
        f"R${a['banca_apos']:,.2f}" if a.get("banca_apos") else "-",
        a.get("notas",""),
    ]

# ── AÇÃO: NOVA APOSTA ─────────────────────────────────────────
def nova_aposta(wb, aposta):
    ws = wb["Sinais Completos"]

    # Encontra última linha com dados
    max_row = ws.max_row
    # Insere linha após o cabeçalho (linha 1), empurrando as demais
    ws.insert_rows(2)

    vals = linha_aposta(aposta)
    bg_novo = CINZA  # primeira linha sempre começa cinza

    for j, v in enumerate(vals):
        c = ws.cell(row=2, column=j+1, value=v)
        c.font = Font(size=10, name="Calibri")
        c.border = bf()
        c.alignment = Alignment(
            horizontal="left" if j in [3,4,21] else "center",
            vertical="center"
        )

        # Coloração especial
        if j == 18:  # Resultado
            bg, fg = cor_resultado(str(v))
            c.fill = F(bg)
            c.font = Font(bold=True, color=fg, size=10, name="Calibri")
        elif j == 9 and v != "-":  # CLV
            clv_num = float(str(v).replace("%","").replace("+",""))
            c.fill = F(VERDE_CLR if clv_num >= 0 else VERM_CLR)
            c.font = Font(
                color=VERDE_TXT if clv_num >= 0 else VERM_TXT,
                bold=True, size=10, name="Calibri"
            )
        elif j == 12:  # Tier
            tc = {"Elite":"FFD700","Premium":AZUL_CLR,"Padrão":CINZA}
            tf = {"Elite":"000000","Premium":AZUL,"Padrão":"000000"}
            c.fill = F(tc.get(str(v), CINZA))
            c.font = Font(
                bold=str(v)=="Elite",
                color=tf.get(str(v),"000000"),
                size=10, name="Calibri"
            )
        else:
            c.fill = F(BRANCO)  # Nova linha sempre branco

    log(f"Nova aposta inserida: #{aposta.get('id','')} — {aposta.get('jogo','')} [{aposta.get('mercado','')}]")

# ── AÇÃO: RESULTADO ───────────────────────────────────────────
def atualizar_resultado(wb, aposta):
    ws = wb["Sinais Completos"]
    sinal_id = str(aposta.get("id",""))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if str(row[0].value) == sinal_id:
            r = row[0].row

            # Odd fechamento (col I = 9)
            if aposta.get("odd_fechamento"):
                ws.cell(row=r, column=9).value = aposta["odd_fechamento"]

            # CLV (col J = 10)
            if aposta.get("odd_fechamento") and aposta.get("odd_entrada"):
                clv = ((aposta["odd_fechamento"]/aposta["odd_entrada"])-1)*100
                c_clv = ws.cell(row=r, column=10)
                c_clv.value = f"{clv:+.2f}%"
                c_clv.fill = F(VERDE_CLR if clv>=0 else VERM_CLR)
                c_clv.font = Font(
                    color=VERDE_TXT if clv>=0 else VERM_TXT,
                    bold=True, size=10, name="Calibri"
                )

            # Resultado (col S = 19)
            resultado = aposta.get("resultado","")
            if resultado:
                c_res = ws.cell(row=r, column=19)
                bg, fg = cor_resultado(resultado)
                c_res.value = resultado
                c_res.fill = F(bg)
                c_res.font = Font(bold=True, color=fg, size=10, name="Calibri")

            # Retorno (col T = 20)
            if aposta.get("retorno") is not None:
                ws.cell(row=r, column=20).value = f"{aposta['retorno']:+.2f}u"

            # Banca após (col U = 21)
            if aposta.get("banca_apos"):
                ws.cell(row=r, column=21).value = f"R${aposta['banca_apos']:,.2f}"

            log(f"Resultado atualizado: #{sinal_id} — {resultado} "
                f"{'R$'+str(aposta.get('retorno','')) if aposta.get('retorno') else ''}")
            break

    # Atualiza histórico de banca
    _atualizar_historico_banca(wb, aposta)

# ── ATUALIZA HISTÓRICO DE BANCA ───────────────────────────────
def _atualizar_historico_banca(wb, aposta):
    ws = wb["Gestão de Banca"]
    if not aposta.get("banca_apos") or not aposta.get("retorno"):
        return

    hoje = str(aposta.get("data", date.today()))
    # Busca linha do dia
    for row in ws.iter_rows(min_row=30, max_row=ws.max_row):
        if str(row[0].value) == hoje:
            # Atualiza resultado do dia
            retorno_atual = row[3].value or "0u"
            try:
                retorno_num = float(str(retorno_atual).replace("u","").replace("+",""))
            except Exception:
                retorno_num = 0
            novo_retorno = retorno_num + aposta["retorno"]
            row[3].value = f"{novo_retorno:+.2f}u"
            row[4].value = f"R${aposta['banca_apos']:,.2f}"
            return

    # Não achou o dia — insere nova linha
    r = ws.max_row + 1
    banca_inicio = aposta.get("banca_apos", 1000) - aposta.get("retorno", 0)
    vals = [
        hoje,
        f"R${banca_inicio:,.2f}",
        1,
        f"{aposta.get('retorno',0):+.2f}u",
        f"R${aposta.get('banca_apos',0):,.2f}",
        "",
        ""
    ]
    bg = VERDE_CLR if aposta.get("retorno",0) >= 0 else VERM_CLR
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=j+1, value=v)
        c.font = Font(size=10, name="Calibri")
        c.fill = F(bg if j in [3,4] else CINZA)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bf()

    log(f"Histórico banca: dia {hoje} adicionado — R${aposta.get('banca_apos',0):,.2f}")

# ── AÇÃO: ATUALIZAR ODDS (MONITORAMENTO 72H) ──────────────────
def atualizar_odds(wb, aposta):
    ws = wb["Monitoramento 72h"]
    jogo = aposta.get("jogo","")

    for row in ws.iter_rows(min_row=11, max_row=ws.max_row):
        if str(row[0].value) == jogo:
            r = row[0].row
            if aposta.get("odd_entrada"):
                ws.cell(row=r, column=7).value = aposta["odd_entrada"]
            if aposta.get("steam"):
                steam = aposta["steam"]
                # Atualiza status baseado no steam
                if steam < -8:
                    status = "🎯 FECHANDO"
                elif steam < 0:
                    status = "⚡ ATIVO"
                else:
                    status = "👁️ OBSERVAÇÃO"
                ws.cell(row=r, column=6).value = status
            log(f"Odds atualizadas: {jogo}")
            return

    log(f"Jogo não encontrado para atualizar odds: {jogo}")

# ── MAIN ──────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    try:
        payload = json.loads(sys.argv[1])
    except json.JSONDecodeError as e:
        log(f"ERRO JSON: {e}")
        sys.exit(1)

    acao = payload.get("acao","")
    aposta = payload.get("aposta", {})

    if not os.path.exists(EXCEL_PATH):
        log(f"ERRO: arquivo {EXCEL_PATH} não encontrado. "
            "Execute primeiro: python gerar_xlsx.py")
        sys.exit(1)

    log(f"Iniciando ação: {acao}")

    try:
        wb = load_workbook(EXCEL_PATH)

        if acao == "nova_aposta":
            if not aposta:
                log("ERRO: campo 'aposta' obrigatório para nova_aposta")
                sys.exit(1)
            nova_aposta(wb, aposta)

        elif acao == "resultado":
            if not aposta:
                log("ERRO: campo 'aposta' obrigatório para resultado")
                sys.exit(1)
            atualizar_resultado(wb, aposta)

        elif acao == "atualizar_odds":
            if not aposta:
                log("ERRO: campo 'aposta' obrigatório para atualizar_odds")
                sys.exit(1)
            atualizar_odds(wb, aposta)

        elif acao == "full_refresh":
            log("full_refresh: regenerando arquivo completo...")
            wb.close()
            os.system("python gerar_xlsx.py")
            log("full_refresh concluído")
            return

        else:
            log(f"ERRO: ação desconhecida '{acao}'")
            sys.exit(1)

        wb.save(EXCEL_PATH)
        log(f"Arquivo salvo: {EXCEL_PATH}")

    except Exception as e:
        log(f"ERRO ao processar: {e}")
        raise

if __name__ == "__main__":
    main()
